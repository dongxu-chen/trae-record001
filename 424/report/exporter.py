"""
报表导出模块
支持Excel和CSV格式导出
"""
import csv
import io
from datetime import datetime, timedelta
from typing import Optional, List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from loguru import logger

from database.models import Product, PriceHistory, Alert, Promotion


class ReportExporter:
    @staticmethod
    def export_products(source=None, category=None, format='xlsx'):
        products, total = Product.get_all(source=source, category=category, page_size=9999)
        if not products:
            logger.warning("没有可导出的商品数据")
            return None

        data = []
        for p in products:
            data.append({
                '商品ID': p.get('product_id', ''),
                '商品名称': p.get('name', ''),
                '品牌': p.get('brand', ''),
                '分类': p.get('category', ''),
                '来源': p.get('source', ''),
                '当前价格': p.get('current_price', ''),
                '原价': p.get('original_price', ''),
                '是否促销': '是' if p.get('is_on_promotion') else '否',
                '库存状态': '有货' if p.get('in_stock') else '缺货',
                '评分': p.get('rating', ''),
                '评论数': p.get('review_count', ''),
                '链接': p.get('url', ''),
                '更新时间': p.get('updated_at', '').strftime('%Y-%m-%d %H:%M:%S') if isinstance(p.get('updated_at'), datetime) else '',
            })

        df = pd.DataFrame(data)
        return ReportExporter._export(df, f'products_{datetime.now().strftime("%Y%m%d_%H%M%S")}', format)

    @staticmethod
    def export_price_history(product_id, start_date=None, end_date=None, format='xlsx'):
        if start_date is None:
            start_date = datetime.utcnow() - timedelta(days=30)
        if end_date is None:
            end_date = datetime.utcnow()

        history = PriceHistory.get_history(product_id, start_date, end_date)
        if not history:
            logger.warning(f"商品 {product_id} 没有价格历史数据")
            return None

        product = Product.get_by_id(product_id)
        product_name = product.get('name', product_id) if product else product_id

        data = []
        for h in history:
            data.append({
                '商品ID': product_id,
                '商品名称': product_name,
                '价格': h.get('price', ''),
                '原价': h.get('original_price', ''),
                '币种': h.get('currency', ''),
                '库存': '有货' if h.get('in_stock') else '缺货',
                '促销中': '是' if h.get('is_on_promotion') else '否',
                '促销信息': h.get('promotion_info', ''),
                '记录时间': h.get('timestamp', '').strftime('%Y-%m-%d %H:%M:%S') if isinstance(h.get('timestamp'), datetime) else '',
            })

        df = pd.DataFrame(data)
        return ReportExporter._export(df, f'price_history_{product_id}_{datetime.now().strftime("%Y%m%d")}', format)

    @staticmethod
    def export_alerts(alert_type=None, status=None, format='xlsx'):
        alerts, total = Alert.get_alerts(alert_type=alert_type, status=status, page_size=9999)
        if not alerts:
            logger.warning("没有可导出的告警数据")
            return None

        type_names = {
            'price_drop': '价格下跌',
            'price_rise': '价格上涨',
            'stock_out': '商品缺货',
            'promotion': '促销活动',
            'new_product': '新品上架',
        }

        data = []
        for a in alerts:
            data.append({
                '告警类型': type_names.get(a.get('alert_type', ''), a.get('alert_type', '')),
                '商品名称': a.get('product_name', ''),
                '来源': a.get('source', ''),
                '原价': a.get('old_price', ''),
                '现价': a.get('new_price', ''),
                '变动幅度': f"{a.get('change_ratio', 0) * 100:.1f}%" if a.get('change_ratio') else '',
                '状态': '已读' if a.get('status') == 'read' else '未读',
                '详情': a.get('message', ''),
                '告警时间': a.get('created_at', '').strftime('%Y-%m-%d %H:%M:%S') if isinstance(a.get('created_at'), datetime) else '',
            })

        df = pd.DataFrame(data)
        return ReportExporter._export(df, f'alerts_{datetime.now().strftime("%Y%m%d_%H%M%S")}', format)

    @staticmethod
    def export_promotions(source=None, format='xlsx'):
        promotions = Promotion.get_active(source=source)
        if not promotions:
            logger.warning("没有可导出的促销数据")
            return None

        type_names = {
            'discount': '折扣',
            'coupon': '优惠券',
            'flash_sale': '限时秒杀',
            'buy_get': '买赠',
            'full_reduction': '满减',
            'bundle': '套餐',
            'general': '一般促销',
        }

        data = []
        for p in promotions:
            data.append({
                '商品名称': p.get('product_name', ''),
                '促销类型': type_names.get(p.get('promo_type', ''), p.get('promo_type', '')),
                '促销价格': p.get('price', ''),
                '原价': p.get('original_price', ''),
                '促销详情': p.get('promo_info', ''),
                '来源': p.get('source', ''),
                '开始时间': p.get('start_date', '').strftime('%Y-%m-%d %H:%M:%S') if isinstance(p.get('start_date'), datetime) else '',
            })

        df = pd.DataFrame(data)
        return ReportExporter._export(df, f'promotions_{datetime.now().strftime("%Y%m%d")}', format)

    @staticmethod
    def export_comparison_report(products_data: List[dict], format='xlsx'):
        if not products_data:
            return None

        data = []
        for p in products_data:
            data.append({
                '商品名称': p.get('name', ''),
                '我方价格': p.get('our_price', ''),
                '竞品A价格': p.get('competitor_a_price', ''),
                '竞品B价格': p.get('competitor_b_price', ''),
                '最低价格': p.get('lowest_price', ''),
                '最高价格': p.get('highest_price', ''),
                '平均价格': p.get('avg_price', ''),
                '价格差异': p.get('price_diff', ''),
            })

        df = pd.DataFrame(data)
        return ReportExporter._export(df, f'comparison_report_{datetime.now().strftime("%Y%m%d")}', format)

    @staticmethod
    def _export(df, filename, format='xlsx'):
        if format == 'csv':
            return ReportExporter._export_csv(df, filename)
        else:
            return ReportExporter._export_excel(df, filename)

    @staticmethod
    def _export_csv(df, filename):
        output = io.StringIO()
        df.to_csv(output, index=False, encoding='utf-8-sig')
        content = output.getvalue()
        output.close()
        return {
            'filename': f'{filename}.csv',
            'content': content,
            'mimetype': 'text/csv',
        }

    @staticmethod
    def _export_excel(df, filename):
        wb = Workbook()
        ws = wb.active
        ws.title = '数据报表'

        header_font = Font(bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if pd.notna(value) else '')
                cell.alignment = cell_alignment
                cell.border = thin_border

        for col_idx in range(1, len(df.columns) + 1):
            max_length = max(
                len(str(df.columns[col_idx - 1])),
                max((len(str(cell.value)) for cell in ws[2:ws.max_row] if cell.value), default=10)
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = 'A2'

        output = io.BytesIO()
        wb.save(output)
        content = output.getvalue()
        output.close()

        return {
            'filename': f'{filename}.xlsx',
            'content': content,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        }